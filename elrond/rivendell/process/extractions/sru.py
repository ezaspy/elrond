#Original form Source: https://github.com/MarkBaggett/srum-dump/blob/master/srum_dump2.py
from openpyxl.cell import WriteOnlyCell
from Registry import Registry
from datetime import datetime,timedelta
import pyesedb
import struct
import re
import openpyxl
import hashlib
import codecs
import uuid


def BinarySIDtoStringSID(sid_str):
    #Original form Source: https://github.com/google/grr/blob/master/grr/parsers/wmi_parser.py
    """Converts a binary SID to its string representation.
     https://msdn.microsoft.com/en-us/library/windows/desktop/aa379597.aspx
    The byte representation of an SID is as follows:
      Offset  Length  Description
      00      01      revision
      01      01      sub-authority count
      02      06      authority (big endian)
      08      04      subauthority #1 (little endian)
      0b      04      subauthority #2 (little endian)
      ...
    Args:
      sid: A byte array.
    Returns:
      SID in string form.
    Raises:
      ValueError: If the binary SID is malformed.
    """
    if not sid_str:
        return ""
    sid = codecs.decode(sid_str,"hex")
    str_sid_components = [sid[0]]
    # Now decode the 48-byte portion
    if len(sid) >= 8:
        subauthority_count = sid[1]
        identifier_authority = struct.unpack(">H", sid[2:4])[0]
        identifier_authority <<= 32
        identifier_authority |= struct.unpack(">L", sid[4:8])[0]
        str_sid_components.append(identifier_authority)
        start = 8
        for i in range(subauthority_count):
            authority = sid[start:start + 4]
            if not authority:
                break
            if len(authority) < 4:
                raise ValueError("In binary SID '%s', component %d has been truncated. "
                         "Expected 4 bytes, found %d: (%s)",
                         ",".join([str(ord(c)) for c in sid]), i,
                         len(authority), authority)
            str_sid_components.append(struct.unpack("<L", authority)[0])
            start += 4
            sid_str = "S-%s" % ("-".join([str(x) for x in str_sid_components]))
    sid_name = template_lookups.get("Known SIDS",{}).get(sid_str,'unknown')
    return "{} ({})".format(sid_str,sid_name)


def blob_to_string(binblob):
    """Takes in a binary blob hex characters and does its best to convert it to a readable string.
       Works great for UTF-16 LE, UTF-16 BE, ASCII like data. Otherwise return it as hex.
    """
    try:
        chrblob = codecs.decode(binblob,"hex")
    except:
        chrblob = binblob
    try:
        if re.match(b'^(?:[^\x00]\x00)+\x00\x00$', chrblob):
            binblob = chrblob.decode("utf-16-le").strip("\x00")
        elif re.match(b'^(?:\x00[^\x00])+\x00\x00$', chrblob):
            binblob = chrblob.decode("utf-16-be").strip("\x00")
        else:
            binblob = chrblob.decode("latin1").strip("\x00")
    except:
        binblob = "" if not binblob else codecs.decode(binblob,"latin-1")
    return binblob


def ole_timestamp(binblob):
    """converts a hex encoded OLE time stamp to a time string"""
    try:
        td,ts = str(struct.unpack("<d",binblob)[0]).split(".")
        dt = datetime(1899,12,30,0,0,0) + timedelta(days=int(td),seconds=86400 * float("0.{}".format(ts)))
    except:
        dt = "This field is incorrectly identified as an OLE timestamp in the template."
    return dt
 

def file_timestamp(binblob):
    """converts a hex encoded windows file time stamp to a time string"""
    try:
        dt = datetime(1601,1,1,0,0,0) + timedelta(microseconds=binblob/10)
    except:
        dt = "This field is incorrectly identified as a file timestamp in the template"
    return dt


def load_registry_sids(reg_file):
    """Given Software hive find SID usernames"""
    sids = {}
    profile_key = r"Microsoft\Windows NT\CurrentVersion\ProfileList"
    tgt_value = "ProfileImagePath"
    try:
        reg_handle = Registry.Registry(reg_file)
        key_handle = reg_handle.open(profile_key)
        for eachsid in key_handle.subkeys():
            sids_path = eachsid.value(tgt_value).value()
            sids[eachsid.name()] = sids_path.split("\\")[-1]
    except:
        return {}
    return sids


def load_interfaces(reg_file):
    """Loads the names of the wireless networks from the software registry hive"""
    try:
        reg_handle = Registry.Registry(reg_file)
    except:
        return {}
    try:
        int_keys = reg_handle.open('Microsoft\\WlanSvc\\Interfaces')
    except:
        return {}
    profile_lookup = {}
    for eachinterface in int_keys.subkeys():
        if len(eachinterface.subkeys())==0:
            continue
        for eachprofile in eachinterface.subkey("Profiles").subkeys():
            profileid = [x.value() for x in list(eachprofile.values()) if x.name()=="ProfileIndex"][0]
            metadata = list(eachprofile.subkey("MetaData").values())
            for eachvalue in metadata:
                if eachvalue.name()=="Channel Hints":
                    channelhintraw = eachvalue.value()
                    hintlength = struct.unpack("I", channelhintraw[0:4])[0]
                    name = channelhintraw[4:hintlength+4] 
                    profile_lookup[str(profileid)] = name.decode(encoding="latin1")
    return profile_lookup


def load_srumid_lookups(database):
    """loads the SRUMID numbers from the SRUM database"""
    id_lookup = {}
    #Note columns  0 = Type, 1 = Index, 2 = Value
    lookup_table = database.get_table_by_name('SruDbIdMapTable')
    column_lookup = dict([(x.name,index) for index,x in enumerate(lookup_table.columns)]) 
    for rec_entry_num in range(lookup_table.number_of_records):
        bin_blob = smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdBlob'])
        if smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdType'])==3:
            bin_blob = BinarySIDtoStringSID(bin_blob)
        elif not bin_blob == "Empty":
            bin_blob = blob_to_string(bin_blob)
        id_lookup[smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdIndex'])] = bin_blob
    return id_lookup


def load_template_lookups(template_workbook):
    """Load any tabs named lookup-xyz form the template file for lookups of columns with the same format type"""
    template_lookups = {}
    for each_sheet in template_workbook.get_sheet_names():
        if each_sheet.lower().startswith("lookup-"):
            lookupname = each_sheet.split("-")[1]
            template_sheet = template_workbook.get_sheet_by_name(each_sheet)
            lookup_table = {}
            for eachrow in range(1,template_sheet.max_row+1):
                value = template_sheet.cell(row = eachrow, column = 1).value
                description = template_sheet.cell(row = eachrow, column = 2).value
                lookup_table[value] = description
            template_lookups[lookupname] = lookup_table
    return template_lookups
    

def load_template_tables(template_workbook):
    """Load template tabs that define the field names and formats for tables found in SRUM"""
    template_tables = {}    
    sheets = template_workbook.get_sheet_names()
    for each_sheet in sheets:
        #open the first sheet in the template
        template_sheet = template_workbook.get_sheet_by_name(each_sheet)
        #retieve the name of the ESE table to populate the sheet with from A1
        ese_template_table = template_sheet.cell(row=1,column=1).value
        #retrieve the names of the ESE table columns and cell styles from row 2 and format commands from row 3 
        template_field = {}
        #Read the first Row B & C in the template into lists so we know what data we are to extract
        for eachcolumn in range(1,template_sheet.max_column+1):
            field_name = template_sheet.cell(row = 2, column = eachcolumn).value
            if field_name == None:
                break
            template_style = template_sheet.cell(row = 4, column = eachcolumn).style
            template_format = template_sheet.cell(row = 3, column = eachcolumn).value
            template_value = template_sheet.cell(row = 4, column = eachcolumn ).value
            if not template_value:
                template_value= field_name
            template_field[field_name] = (template_style,template_format,template_value)
        template_tables[ese_template_table] = (each_sheet, template_field)
    return template_tables    


def smart_retrieve(ese_table, ese_record_num, column_number):
    """Given a row and column will determine the format and retrieve a value from the ESE table"""
    rec = ese_table.get_record(ese_record_num)
    col_type = rec.get_column_type(column_number)
    col_data = rec.get_value_data(column_number)
    if col_type == pyesedb.column_types.BINARY_DATA:
        col_data = "" if not col_data else codecs.encode(col_data,"HEX")
    elif col_type == pyesedb.column_types.BOOLEAN:
        col_data = struct.unpack('?',col_data)[0]
    elif col_type == pyesedb.column_types.CURRENCY:
        pass
    elif col_type == pyesedb.column_types.DATE_TIME:
        col_data = ole_timestamp(col_data)
    elif col_type == pyesedb.column_types.DOUBLE_64BIT:
        col_data = 0 if not col_data else struct.unpack('d',col_data)[0]
    elif col_type == pyesedb.column_types.FLOAT_32BIT:
        col_data = 0.0 if not col_data else struct.unpack('f',col_data)[0]
    elif col_type == pyesedb.column_types.GUID:
        col_data = 0 if not col_data else str(uuid.UUID(bytes = col_data))
    elif col_type == pyesedb.column_types.INTEGER_16BIT_SIGNED:
        col_data = 0 if not col_data else struct.unpack('h',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_16BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('H',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_32BIT_SIGNED:
        col_data =  0 if not col_data else struct.unpack('i',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_32BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('I',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_64BIT_SIGNED:
        col_data = 0 if not col_data else struct.unpack('q',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_8BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('B',col_data)[0]
    elif col_type == pyesedb.column_types.LARGE_BINARY_DATA:
        col_data = "" if not col_data else codecs.encode(col_data,"HEX")
    elif col_type == pyesedb.column_types.LARGE_TEXT:
        col_data = blob_to_string(col_data)
    elif col_type == pyesedb.column_types.NULL:
        pass
    elif col_type == pyesedb.column_types.SUPER_LARGE_VALUE:
        col_data = "" if not col_data else codecs.encode(col_data,"HEX")
    elif col_type == pyesedb.column_types.TEXT:
        col_data = blob_to_string(col_data)   
    else:
        col_data = blob_to_string(col_data)    
    if col_data==None:
        col_data = "Empty"
    return col_data


def format_output(val, eachformat, eachstyle, xls_sheet):
    """Returns a excel cell with the data formated as specified in the template table"""
    new_cell = WriteOnlyCell(xls_sheet, value = "init")
    new_cell.style = eachstyle
    if val==None:
        val="None"
    elif eachformat in [None, "OLE"]:
        pass
    elif eachformat=="FILE":
        val = file_timestamp(val)
        new_cell.number_format = 'YYYY MMM DD'
    elif eachformat.startswith("FILE:"):
        val = file_timestamp(val)
        val = val.strftime(eachformat[5:])
    elif eachformat.lower().startswith("lookup-"):
        lookup_name = eachformat.split("-")[1]
        if lookup_name in template_lookups:
            lookup_table = template_lookups.get(lookup_name,{})
            val = lookup_table.get(val,val)
    elif eachformat.lower() == "lookup_id":
        val = id_table.get(val, "No match in srum lookup table for %s" % (val))
    elif eachformat.lower() == "lookup_luid":
        inttype = struct.unpack(">H6B", codecs.decode(format(val,'016x'),'hex'))[0]
        val = template_lookups.get("LUID Interfaces",{}).get(inttype,"")
    elif eachformat.lower() == "seconds":
        val = val/86400.0
        new_cell.number_format = 'dd hh:mm:ss'
    elif eachformat.lower() == "md5":
        val = hashlib.md5(str(val)).hexdigest()
    elif eachformat.lower() == "sha1":
        val = hashlib.sha1(str(val)).hexdigest()
    elif eachformat.lower() == "sha256":
        val = hashlib.sha256(str(val)).hexdigest()
    elif eachformat.lower() == "base16":
        if type(val)==int:
            val = hex(val)
        else:
            val = format(val,"08x")
    elif eachformat.lower() == "base2":
        if type(val)==int:
            val = format(val,"032b")
        else:
            try:
                val = int(str(val),2)
            except :
                val = val
    else:
        val = val
    new_cell.value = val
    return new_cell


def process_srum(ese_db, target_wb ):
    """Process all the tables and columns in the ESE database"""
    for table_num in range(ese_db.number_of_tables):
        ese_table = ese_db.get_table(table_num)
        if ese_table.name in skip_tables:
            continue
        if ese_table.name in template_tables:
            tname,tfields = template_tables.get(ese_table.name)
        else:
            tname = ese_table.name[1:15]
        xls_sheet = target_wb.create_sheet(title=tname)
        header_row = [x.name for x in ese_table.columns]
        if ese_table.name in template_tables:
            tname,tfields = template_tables.get(ese_table.name)
            header_row = []
            for eachcol in ese_table.columns:
                if eachcol.name in tfields:
                    cell_style, _, cell_value = tfields.get(eachcol.name)
                    new_cell = WriteOnlyCell(xls_sheet, value=cell_value)
                    new_cell.style = cell_style
                    header_row.append( new_cell )
                else:
                    header_row.append(WriteOnlyCell(xls_sheet, value=eachcol.name))
        xls_sheet.append(header_row)
        column_names = [x.name for x in ese_table.columns]
        for row_num in range(ese_table.number_of_records):
            try:
                ese_row = ese_table.get_record(row_num)
            except:
                continue
            if ese_row == None:
                continue
            xls_row = []
            for col_num in range(ese_table.number_of_columns):
                val = smart_retrieve(ese_table,row_num, col_num)
                if val=="Error":
                    val = "WARNING: Invalid Column Name {}".format(column_names[col_num])
                elif val==None:
                    val="None"  
                elif ese_table.name in template_tables:
                    tname,tfields = template_tables.get(ese_table.name) 
                    if column_names[col_num] in tfields:
                        cstyle, cformat, _ = tfields.get(column_names[col_num])
                        val = format_output(val, cformat, cstyle,xls_sheet)              
                xls_row.append(val)
            xls_sheet.append(xls_row)


srum_path = "<SRUDB.DAT>"
temp_path = "/opt/elrond/elrond/tools/srum-dump/.SRUM_TEMPLATE2.xlsx"

ese_db = pyesedb.file()
ese_db.open("<SRUDB.dat>")
template_wb = openpyxl.load_workbook("/opt/elrond/elrond/tools/srum-dump/.SRUM_TEMPLATE2.xlsx")

skip_tables = ['MSysObjects', 'MSysObjectsShadow', 'MSysObjids', 'MSysLocales','SruDbIdMapTable']
template_tables = load_template_tables(template_wb)
template_lookups = load_template_lookups(template_wb)
id_table = load_srumid_lookups(ese_db)

target_wb = openpyxl.Workbook()
process_srum(ese_db, target_wb)

firstsheet=target_wb.get_sheet_by_name("Sheet")
target_wb.remove_sheet(firstsheet)
target_wb.save("<SRUM_DUMP_OUTPUT.xlsx>")
